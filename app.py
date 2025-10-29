import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io

st.title("ãƒ¡ãƒ‹ãƒ¥ãƒ¼åä¸Šæ›¸ããƒ„ãƒ¼ãƒ«")

uploaded_menu_file = st.file_uploader("ãƒ¡ãƒ‹ãƒ¥ãƒ¼åå¤‰æ›´ä¾é ¼ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ï¼ˆãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰è§£é™¤ã—ã¦ï¼‰", type="xlsx")
uploaded_code_file = st.file_uploader("åª’ä½“ã‚³ãƒ¼ãƒ‰ç™ºç•ªä¾é ¼ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰ï¼ˆãƒ‘ã‚¹ãƒ¯ãƒ¼ãƒ‰è§£é™¤ã—ã¦ï¼‰", type="xlsx")

if uploaded_menu_file and uploaded_code_file:
    menu_wb = load_workbook(uploaded_menu_file)
    code_wb = load_workbook(uploaded_code_file)

    st.subheader("ã‚·ãƒ¼ãƒˆåã®ç¢ºèª")
    menu_sheet_name = st.selectbox("ãƒ¡ãƒ‹ãƒ¥ãƒ¼åå¤‰æ›´ä¾é ¼ãƒ•ã‚¡ã‚¤ãƒ«ã®ã‚·ãƒ¼ãƒˆã‚’é¸æŠ", menu_wb.sheetnames)
    code_sheet_name = st.selectbox("åª’ä½“ã‚³ãƒ¼ãƒ‰ç™ºç•ªä¾é ¼ãƒ•ã‚¡ã‚¤ãƒ«ã®ã‚·ãƒ¼ãƒˆã‚’é¸æŠ", code_wb.sheetnames)

    menu_ws = menu_wb[menu_sheet_name]
    code_ws = code_wb[code_sheet_name]

    # ãƒ‡ãƒ¼ã‚¿æŠ½å‡º
    menu_data = []
    for row in menu_ws.iter_rows(min_row=2, values_only=True):
        media_code = row[7]  # Håˆ—
        menu_name = row[8]   # Iåˆ—
        point = row[10]      # Kåˆ—
        media_name = row[11] # Låˆ—
        menu_data.append((media_code, menu_name, point, media_name))

    # ãƒ—ãƒ¬ãƒ“ãƒ¥ãƒ¼ç”¨ãƒ‡ãƒ¼ã‚¿
    preview_rows = []
    for row in code_ws.iter_rows(min_row=2):
        target_code = row[1].value  # Båˆ—
        for m_code, m_name, m_point, m_media_name in menu_data:
            if target_code == m_code:
                preview_rows.append({
                    "åª’ä½“ã‚³ãƒ¼ãƒ‰": target_code,
                    "æ—§ãƒ¡ãƒ‹ãƒ¥ãƒ¼å": row[2].value,
                    "æ–°ãƒ¡ãƒ‹ãƒ¥ãƒ¼å": m_name,
                    "æ—§ãƒã‚¤ãƒ³ãƒˆ": row[4].value,
                    "æ–°ãƒã‚¤ãƒ³ãƒˆ": m_point,
                    "æ—§åª’ä½“å": row[5].value,
                    "æ–°åª’ä½“å": m_media_name
                })
                break

    if preview_rows:
        st.subheader("æ›´æ–°ãƒ—ãƒ¬ãƒ“ãƒ¥ãƒ¼")
        st.dataframe(pd.DataFrame(preview_rows))

    if st.button("ä¸Šæ›¸ãå®Ÿè¡Œ"):
        rose_fill = PatternFill(start_color="FF66CC", end_color="FF66CC", fill_type="solid")
        updated_count = 0
        for row in code_ws.iter_rows(min_row=2):
            target_code = row[1].value  # Båˆ—
            for m_code, m_name, m_point, m_media_name in menu_data:
                if target_code == m_code:
                    row[2].value = m_name       # Cåˆ—
                    row[2].fill = rose_fill
                    row[4].value = m_point      # Eåˆ—
                    row[4].fill = rose_fill
                    row[5].value = m_media_name # Fåˆ—
                    row[5].fill = rose_fill
                    updated_count += 1
                    break

        st.success(f"{updated_count} ä»¶ã®ã‚»ãƒ«ã‚’æ›´æ–°âœ¨æ›´æ–°ã‚»ãƒ«ã¯ãƒ”ãƒ³ã‚¯ã§ã™ğŸ¦©ã‚¨ã‚¯ã‚¹ãƒãƒ¼ãƒˆã—ã¦ç¢ºèªã—ã¦ãã ã•ã„ï¼")

        output = io.BytesIO()
        code_wb.save(output)
        st.download_button(
            label="æ›´æ–°æ¸ˆã¿ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰",
            data=output.getvalue(),
            file_name="æ›´æ–°æ¸ˆã¿_åª’ä½“ã‚³ãƒ¼ãƒ‰ç™ºç•ªä¾é ¼.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )